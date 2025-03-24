from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import sys

def adjust_col(ws, larguras_minimas):
    for col_idx, min_width in larguras_minimas.items():
        max_length = min_width  # Começa com a largura mínima
        for cell in ws[col_idx]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_idx].width = max_length + 2  # Adiciona um pequeno buffer

def format(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="16365c", end_color="4F81BD", fill_type="solid")
    cell_border = Border(left=Side(style="thin"), right=Side(style="thin"), 
                         top=Side(style="thin"), bottom=Side(style="thin"))
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Definir larguras mínimas e ajustar dinamicamente
    larguras_minimas = {"A": 10, "B": 27, "C": 25, "D": 25, "E": 10, "F": 20}
    adjust_col(ws, larguras_minimas)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = cell_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="b8d9ff", end_color="b8d9ff", fill_type="solid")

    wb.save(excel_path)
    print("Tabela formatada com sucesso!")

excel_path = sys.argv[1]
format(excel_path)