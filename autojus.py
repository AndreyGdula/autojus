import fitz
import pandas as pd
import re
import os
import sys
import subprocess as sb
from openpyxl import load_workbook


# Função para extrair texto do PDF
def extrair_texto_pdf(pdf_path):
    texto = ""
    with fitz.open(pdf_path) as doc:
        for page in doc:
            texto += page.get_text("text") + "\n"
    return texto

def extrair_dados_processos(pdf_path, padrao_processo, padrao_autor, padrao_advogado, padrao_oab, padrao_data, confirm_callback=None, message_callback=None):
    texto_extraido = extrair_texto_pdf(pdf_path)

    # Encontrando todas as ocorrências
    arquivo = os.path.splitext(os.path.basename(pdf_path))[0]
    processos = re.findall(padrao_processo, texto_extraido)
    autores = re.findall(padrao_autor, texto_extraido)
    advogados = re.findall(padrao_advogado, texto_extraido)
    oab = re.findall(padrao_oab, texto_extraido)
    datas = re.findall(padrao_data, texto_extraido)

    # Verificar se algum padrão não foi encontrado
    padroes_nao_encontrados = []
    if not processos:
        padroes_nao_encontrados.append("Número do Processo")
    if not autores:
        padroes_nao_encontrados.append("Autor")
    if not advogados:
        padroes_nao_encontrados.append("Advogado")
    if not oab:
        padroes_nao_encontrados.append("OAB")
    if not datas:
        padroes_nao_encontrados.append("Data de Distribuição")

    if padroes_nao_encontrados:
        continuar = confirm_callback(f"Os seguintes padrões não foram encontrados no arquivo PDF: {', '.join(padroes_nao_encontrados)}.\nDeseja continuar com as informações encontradas?")
        if not continuar:
            message_callback("Operação cancelada.")
            return []

    # Preencher valores em branco para padrões não encontrados
    if not processos:
        processos = [""]
    if not autores:
        autores = [""]
    if not advogados:
        advogados = [""]
    if not oab:
        oab = [""]
    if not datas:
        datas = [""]

    # Organizando os dados em uma lista
    dados = []
    for i in range(len(processos)):
        dados.append([arquivo, processos[i], autores[i], advogados[i], oab[i], datas[i]])

    return dados

# Garantir que a coluna Arquivo fique na coluna A
def move_col(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    # Encontrar a coluna "Arquivo" e movê-la para a posição A
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value == "Arquivo":
            ws.move_range(f"{cell.column_letter}1:{cell.column_letter}{ws.max_row}", rows=0, cols=1 - col_idx)
            break

    wb.save(excel_path)
    wb.close()


def main(pdf_path, confirm_callback, message_callback):
    # Regex
    padrao_processo = r"Processo nº: (\d{7}-\d{2}\.\d{4}\.\d\.\d{2}\.\d{4})"
    padrao_autor = r"Autor: (.+)"
    padrao_advogado = r"Advogado: ([^-]+)"
    padrao_oab = r"OAB[:\s]+(\d+)"
    padrao_data = r"Data de Distribuição: (\d{1,2}/\d{1,2}/\d{2,4})"

    excel_path = "output/processos_extraidos.xlsx"  # Caminho do Excel
    dados_processos = extrair_dados_processos(pdf_path, padrao_processo, padrao_autor, padrao_advogado, padrao_oab, padrao_data, confirm_callback, message_callback)  # Extraindo os dados

    # Criando um DataFrame com os novos dados
    df_novo = pd.DataFrame(dados_processos, columns=["Arquivo", "Número do Processo", "Autor", "Advogado", "OAB", "Data de Distribuição"])

    # Verificando se o arquivo Excel já existe
    if os.path.exists(excel_path):
        df_existente = pd.read_excel(excel_path)
        if df_existente["Número do Processo"].str.contains(df_novo["Número do Processo"].iloc[0]).any():
            confirm_edit = confirm_callback(f"O processo {df_novo['Número do Processo'].iloc[0]} já existe no arquivo Excel. Deseja atualizar as informações?")
            if confirm_edit:
                df_final = pd.concat([df_existente, df_novo]).drop_duplicates(subset=["Número do Processo"], keep="last")
                message_callback("Processo atualizado com sucesso.")
            else:
                message_callback("Operação cancelada.")
                df_final = df_existente
        else:
            df_final = pd.concat([df_existente, df_novo]).drop_duplicates(subset=["Número do Processo"], keep="last")
            message_callback(f"Processos extraídos em {excel_path}")
    else:
        df_final = df_novo

    # Salvando no Excel
    df_final.to_excel(excel_path, index=False)
    move_col(excel_path)
    sb.run(["python", "format_table.py", excel_path])  # Formatar arquivo Excel
